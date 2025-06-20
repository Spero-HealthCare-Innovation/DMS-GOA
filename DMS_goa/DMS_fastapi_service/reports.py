from fastapi import APIRouter, Query
from typing import Optional
from DMS_goa.hive_connector import hive_connecter_execution
from io import BytesIO
import pandas as pd
from admin_web.models import *
from fastapi.responses import StreamingResponse
from datetime import datetime, timedelta
import pytz
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

india_tz = pytz.timezone('Asia/Kolkata')


router = APIRouter()

@router.get('/incident_report_incident_daywise')
def incident_report_incident_daywise(
    from_date : Optional[str] = Query(..., description="Start date in YYYY-MM-DD format"),
    to_date : Optional[str] = Query(..., description="End date in YYYY-MM-DD format")):
    try:
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
        to_date_plus_one = to_date_obj.strftime("%Y-%m-%d")

        incident_data = DMS_Incident.objects.filter(inc_added_date__range = (from_date,to_date_plus_one), inc_type=1)
        # data=[]
        # for i in incident_data:
        #     bb = []
        #     noti=DMS_Notify.objects.filter(incident_id=i,not_is_deleted=False)
        #     for j in noti:
        #         for k in j.alert_type_id:
        #             res=DMS_Responder.objects.get(responder_id=int(k))
        #             bb.append(res.responder_name)
        #     dt = {
        #     "incident_id": str(i.incident_id),
        #     "incident_datetime":i.inc_datetime,
        #     "disaster_name": i.disaster_type.disaster_name if i.disaster_type else None,
        #     "incident_type": "Emergency" if i.inc_type==1 else "Non-Emergency",
        #     "alert_type": "High" if i.alert_type==1 else "Medium" if i.alert_type==2 else "Low" if i.alert_type==3 else "Very Low" if i.alert_type==4 else "",
        #     "responder": list(set(bb))
        #     }
        #     data.append(dt)

        data = [{
            "incident_id": str(incident.incident_id),
            "incident_datetime": incident.inc_datetime,
            "disaster_name": incident.disaster_type.disaster_name if incident.disaster_type else None,
            "incident_type": "Emergency",
            "alert_type": ("High" if incident.alert_type == 1 else "Medium" if incident.alert_type == 2 else "Low" if incident.alert_type == 3 else "Very Low" if incident.alert_type == 4 else "" ),
            "responder": list({DMS_Responder.objects.get(responder_id=int(k)).responder_name for notif in DMS_Notify.objects.filter(incident_id=incident, not_is_deleted=False) for k in notif.alert_type_id })
        } for incident in incident_data]
        return data
    except Exception as e:
        return {"Error":"Error","msg":str(e)}



@router.get('/download_incident_report_incident_daywise')
def download_incident_report_incident_daywise(
    from_date: Optional[str] = Query(..., description="Start date in YYYY-MM-DD format"),
    to_date: Optional[str] = Query(..., description="End date in YYYY-MM-DD format")
):
    try:
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
        to_date_plus_one = to_date_obj.strftime("%Y-%m-%d")

        incident_data = DMS_Incident.objects.filter(
            inc_added_date__range=(from_date, to_date_plus_one)
        )

        # Excel workbook setup
        wb = Workbook()
        ws = wb.active
        ws.title = "Incident Report"

        headers = [
            "incident_id", "incident_datetime", "disaster_name", "incident_type", "alert_type", "responder",
            "Caller Number", "Caller Name", "Call Duration", "Incident Address", "Incident District",
            "Incident Tahsil", "Ward", "Incident Summary", "Remark"
        ]
        ws.append(headers)

        row_num = 2  # Data starts from row 2

        for i in incident_data:
            # Get responders
            responders = list({
                DMS_Responder.objects.get(responder_id=int(k)).responder_name
                for j in DMS_Notify.objects.filter(incident_id=i, not_is_deleted=False)
                for k in j.alert_type_id
            })

            # Get remarks
            remark_objs = DMS_Comments.objects.filter(incident_id=i, comm_is_deleted=False)
            remark_list = [z.comments for z in remark_objs] or [""]

            start_row = row_num

            for remark in remark_list:
                ws.append([
                    str(i.incident_id),
                    i.inc_datetime,
                    i.disaster_type.disaster_name if i.disaster_type else None,
                    "Emergency" if i.inc_type == 1 else "Non-Emergency",
                    "High" if i.alert_type == 1 else "Medium" if i.alert_type == 2 else
                    "Low" if i.alert_type == 3 else "Very Low" if i.alert_type == 4 else None,
                    ', '.join(responders),
                    i.caller_id.caller_no if i.caller_id else None,
                    i.caller_id.caller_name if i.caller_id else None,
                    i.time,
                    i.location,
                    i.district.dis_name if i.district else None,
                    i.tahsil.tah_name if i.tahsil else None,
                    i.ward.ward_name if i.ward else None,
                    i.summary.summary if i.summary else None,
                    remark
                ])
                row_num += 1

            end_row = row_num - 1

            # Merge all cells except the last column ("Remark") if more than 1 remark
            if end_row > start_row:
                for col_index in range(1, len(headers)):  # Skip last column
                    cell_range = f"{get_column_letter(col_index)}{start_row}:{get_column_letter(col_index)}{end_row}"
                    ws.merge_cells(cell_range)

        # Save Excel to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"incident_report_{from_date}_to_{to_date}.xlsx"
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )

    except Exception as e:
        return {"Error": "Error", "msg": str(e)}

# def download_incident_report_incident_daywise(
#     from_date:Optional[str]=Query(...,description="Start date in YYYY-MM-DD format"),
#     to_date:Optional[str]=Query(...,description="End date in YYYY-MM-DD format")):
#     try:
#         to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
#         to_date_plus_one = to_date_obj.strftime("%Y-%m-%d")

#         incident_data = DMS_Incident.objects.filter(inc_added_date__range = (from_date,to_date_plus_one))
#         # data=[{"incident_id": str(i.incident_id),"disaster_type": i.disaster_type.disaster_name if i.disaster_type else None,"incident_datetime": i.inc_datetime,"clouser_status": i.clouser_status,"incident_remark": i.comment_id.comments if i.comment_id else None} for i in incident_data]

#         # data = [{
#         #     "incident_id": str(incident.incident_id),
#         #     "incident_datetime": incident.inc_datetime,
#         #     "disaster_name": incident.disaster_type.disaster_name if incident.disaster_type else None,
#         #     "incident_type": "Emergency",
#         #     "alert_type": ("High" if incident.alert_type == 1 else "Medium" if incident.alert_type == 2 else "Low" if incident.alert_type == 3 else "Very Low" if incident.alert_type == 4 else "" ),
#         #     "responder": list({DMS_Responder.objects.get(responder_id=int(k)).responder_name for notif in DMS_Notify.objects.filter(incident_id=incident, not_is_deleted=False) for k in notif.alert_type_id })
#         # } for incident in incident_data]



#         data=[]
#         for i in incident_data:
#             bb = []
#             noti=DMS_Notify.objects.filter(incident_id=i,not_is_deleted=False)
#             for j in noti:
#                 for k in j.alert_type_id:
#                     res=DMS_Responder.objects.get(responder_id=int(k))
#                     bb.append(res.responder_name)
#             gg=DMS_Comments.objects.filter(incident_id=i,comm_is_deleted=False)
#             uu=[z.comments for z in gg]
#             dt = {
#             "incident_id": str(i.incident_id),
#             "incident_datetime":i.inc_datetime,
#             "disaster_name": i.disaster_type.disaster_name if i.disaster_type else None,
#             "incident_type": "Emergency" if i.inc_type==1 else "Non-Emergency",
#             "alert_type": "High" if i.alert_type==1 else "Medium" if i.alert_type==2 else "Low" if i.alert_type==3 else "Very Low" if i.alert_type==4 else None,
#             "responder": list(set(bb)),
#             "Caller Number": i.caller_id.caller_no if i.caller_id else None,
#             "Caller Name":i.caller_id.caller_name if i.caller_id else None,
#             "Call Duration":i.time,
#             "Incident Address": i.location,
#             "Incident District":i.district.dis_name if i.district else None,
#             "Incident Tahsil":i.tahsil.tah_name if i.tahsil else None,
#             "Ward": i.ward.ward_name if i.ward else None,
#             "Incident Summary":i.summary.summary if i.summary else  None,
#             "Remark":uu
#             }
#             data.append(dt)


#         df = pd.DataFrame(data)
#         output = BytesIO()
#         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#             df.to_excel(writer, index=False, sheet_name='Incident Report')
#         output.seek(0)
#         filename = f"incident_report_{from_date}_to_{to_date}.xlsx"
#         headers = {
#             'Content-Disposition': f'attachment; filename="{filename}"'
#         }
#         return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)
#     except Exception as e:
#         return {"Error":'Error',"msg":str(e)}









@router.get("/incident_closure_report_daywise/")
def incident_report_daywise(
    from_date: Optional[str] = Query(..., description="Start date in YYYY-MM-DD format"),
    to_date: Optional[str] = Query(..., description="End date in YYYY-MM-DD format")):
    try:
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
        to_date_plus_one = to_date_obj.strftime("%Y-%m-%d")

        closure_data = DMS_incident_closure.objects.filter(closure_added_date__range=(from_date, to_date_plus_one))
        dt = []
        for i in closure_data:
            nn={
                "incident_id": str(i.incident_id.incident_id if i.incident_id else None),
                "disaster_type": i.incident_id.disaster_type.disaster_name if i.incident_id and i.incident_id.disaster_type else None,
                "closure_acknowledge": i.closure_acknowledge,
                "closure_start_base_location": i.closure_start_base_location,
                "closure_at_scene": i.closure_at_scene,
                "closure_from_scene": i.closure_from_scene,
                "closure_back_to_base": i.closure_back_to_base,
                "closure_remark": i.closure_remark
            }
            dt.append(nn)
        return dt
    except Exception as e:
        return {"status": "error","message": str(e)}





@router.get("/download_incident_closure_report_daywise/")
def incident_report_daywise(
    from_date: Optional[str] = Query(..., description="Start date in YYYY-MM-DD format"),
    to_date: Optional[str] = Query(..., description="End date in YYYY-MM-DD format")
):
    try:
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
        to_date_plus_one = to_date_obj.strftime("%Y-%m-%d")
        closure_data = DMS_incident_closure.objects.filter(closure_added_date__range=(from_date, to_date_plus_one))
        dt = []
        for i in closure_data:
            if i.incident_id:
                nn={
                    "incident_id": str(i.incident_id.incident_id if i.incident_id else None),
                    "Alert Source":"System Alert" if  i.incident_id.mode == 2 else "Manual Calls",
                    "disaster_type": i.incident_id.disaster_type.disaster_name if i.incident_id and i.incident_id.disaster_type else None,
                    "Alert Type": "High" if i.incident_id.alert_type == 1 else "Medium" if i.incident_id.alert_type == 2 else "Low" if i.incident_id.alert_type == 3 else "Very Low" if i.incident_id.alert_type == 4 else "Unknown",
                    "Alert Time" : i.incident_id.alert_id.alert_datetime if i.incident_id and i.incident_id.alert_id else None,
                    "Incident Dispatch Time" : i.incident_id.inc_added_date,
                    "Closure Time": i.closure_added_date,
                    "closure_acknowledge": i.closure_acknowledge,
                    "closure_start_base_location": i.closure_start_base_location,
                    "closure_at_scene": i.closure_at_scene,
                    "closure_from_scene": i.closure_from_scene,
                    "closure_back_to_base": i.closure_back_to_base,
                    "closure_remark": i.closure_remark,
                    "Caller Number" : i.incident_id.caller_id.caller_name if i.incident_id and i.incident_id.caller_id else None,
                    "Caller Name" : i.incident_id.caller_id.caller_no if i.incident_id and i.incident_id.caller_id else None,
                    "Address" : i.incident_id.location if i.incident_id else None,
                    "Lattitude" : i.incident_id.latitude if i.incident_id else None,
                    "Longitude" : i.incident_id.longitude if i.incident_id else None
                }

            
            dt.append(nn)

        df = pd.DataFrame(dt)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Incident Report')
        output.seek(0)

        filename = f"incident_report_{from_date}_to_{to_date}.xlsx"
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }

        return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)

    except Exception as e:
        return {"status": "error", "message": str(e)}



# Closure report from Hive db

# @router.get("/download_incident_closure_report_daywise/")
# def incident_report_daywise(
#     from_date: Optional[str] = Query(..., description="Start date in YYYY-MM-DD format"),
#     to_date: Optional[str] = Query(..., description="End date in YYYY-MM-DD format")
# ):
#     try:
#         to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
#         to_date_plus_one = to_date_obj.strftime("%Y-%m-%d")
#         query = f"""
#             SELECT incident_id, disaster_type_id, alert_id_id, inc_added_date, mode, alert_type,caller_id_id, location, latitude, longitude, closure_acknowledge, closure_start_base_location, 
#                    closure_at_scene, closure_from_scene, closure_back_to_base, closure_remark, closure_added_date 
#             FROM final_closure_report 
#             WHERE closure_added_date BETWEEN '{from_date}' AND '{to_date_plus_one}'
#         """
#         data = hive_connecter_execution(query)
#         dt=[] 
#         for i in data:
#             dstss_dt = DMS_Disaster_Type.objects.filter(disaster_id=i['disaster_type_id'])
#             if dstss_dt.exists():
#                 dstss = dstss_dt.first()
#             else:
#                 dstss = None
#             print(dstss,'dstssdstssdstssdstss')
#             # alrt = Weather_alerts.objects.get(pk_id=i['alert_id_id'] if i['alert_id_id'] else 0)
#             alrt_qs = Weather_alerts.objects.filter(pk_id=i['alert_id_id'])
#             if alrt_qs.exists():
#                 alrt = alrt_qs.first()
#             else:
#                 alrt = None  
            
#             caller_dtls = DMS_Caller.objects.filter(caller_pk_id=i['caller_id_id'])
#             print(caller_dtls,'caller_dtlscaller_dtlscaller_dtls')
#             if caller_dtls.exists():
#                 caller=caller_dtls.first()
#             else:
#                 caller=None
            
#             nn={"Incident Id": i['incident_id'],
#                 "Alert Source":"System Alert" if  i['mode'] == 2 else "Manual Calls",
#                 "Disaster Type": dstss.disaster_name if dstss else None,
#                 "Alert Type": "High" if i['alert_type'] == 1 else "Medium" if i['alert_type'] == 2 else "Low" if i['alert_type'] == 3 else "Very Low" if i['alert_type'] == 4 else "Unknown",
#                 "Alert Time" : alrt.alert_datetime.astimezone(india_tz).replace(tzinfo=None) if alrt else "",
#                 "Incident Dispatch Time" : i['inc_added_date'],
#                 "Closure Time": i['closure_added_date'],
#                 "Closure Acknowledge" : i['closure_acknowledge'],
#                 "Closure Start Base location" : i['closure_start_base_location'],
#                 "Closure At Scene" : i['closure_at_scene'],
#                 "Closure From Scene" : i['closure_from_scene'],
#                 "Closure Back To Base" : i['closure_back_to_base'],
#                 "Closure Remark" : i['closure_remark'],
#                 "Caller Number" : caller.caller_no if caller else None,
#                 "Caller Name" : caller.caller_name if caller else None,
#                 "Address" : i['location'],
#                 "Lattitude" : i['latitude'],
#                 "Longitude" : i['longitude']
#             }
#             dt.append(nn)


#         df = pd.DataFrame(dt)
#         output = BytesIO()
#         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#             df.to_excel(writer, index=False, sheet_name='Incident Report')
#         output.seek(0)

#         filename = f"incident_report_{from_date}_to_{to_date}.xlsx"
#         headers = {
#             'Content-Disposition': f'attachment; filename="{filename}"'
#         }

#         return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)

#     except Exception as e:
#         return {"status": "error", "message": str(e)}


















